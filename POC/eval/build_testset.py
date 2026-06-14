"""
Build eval/testset.jsonl from the volunteer-annotated question set.

Source (outside the repo, gitignored): docs/Austria/Test Qs.xlsx — 13 questions
from a domain expert (Анастасия), columns: Вопрос / Ответ / Примечание (gold §§)
/ Сложность (easy|difficult). This script reads that sheet and applies the
reviewed § -> chunk_id mapping to emit one JSON object per question.

Mapping rationale (per question, reviewed with the user):
  - gold_label        — the expert's "Примечание" verbatim.
  - gold_chunks        — precise §-level anchors that exist in the corpus.
  - gold_chunks_secondary — chapter-level "also acceptable" chunks (q07: the
                       whole 8. Hauptstück FPG = deportation chapter).
  - in_corpus / expected_behavior — "вне корпуса" / a regulation we do not hold
                       (EMRK) => refuse; otherwise answer.
  - reference_answer   — the expert's answer (column Ответ), used by the judges.

The testset.jsonl is the committed, reproducible artifact; this builder documents
how it was derived. It needs the source xlsx (not shipped) to regenerate.

    ~/venvs/ambassy-poc/bin/python eval/build_testset.py
"""
import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
SRC = (HERE / ".." / ".." / ".." / "docs" / "Austria" / "Test Qs.xlsx").resolve()
OUT = HERE / "testset.jsonl"

# FPG "8. Hauptstück" (Aufenthaltsbeendende Maßnahmen / deportation) — the §§ that
# fall under it, resolved from the cleaned FPG.md. q07's gold names this chapter.
FPG_8_HAUPTSTUECK = [
    "FPG-52", "FPG-52a", "FPG-53", "FPG-55", "FPG-56", "FPG-57", "FPG-58",
    "FPG-59", "FPG-60", "FPG-61", "FPG-66", "FPG-67", "FPG-69", "FPG-70",
    "FPG-71", "FPG-76", "FPG-77", "FPG-78", "FPG-79", "FPG-80", "FPG-81",
]

# Per-question mapping, keyed by source row order (1..13 == xlsx rows 2..14).
# gold_chunks: precise §-level anchors present in the corpus.
MAPPING = {
    1:  dict(gold_chunks=[],                      secondary=[], notes="статистика МВД — не нормативный текст, вне корпуса"),
    2:  dict(gold_chunks=["AsylG-27a", "AsylG-22"],secondary=[], notes="Настя 2026-06-13: статутные сроки решения в корпусе — §27a (ускоренное, 5 мес.), §22 Abs.6 (Schubhaft, 3 мес.); эмпирические 'до 5 лет' остаются практикой"),
    3:  dict(gold_chunks=["AuslBG-4"],            secondary=[], notes=""),
    4:  dict(gold_chunks=["AsylG-13", "AsylG-51"],secondary=[], notes=""),
    5:  dict(gold_chunks=["FPG-50", "FPG-45a"],   secondary=[], notes="Настя 2026-06-13: правовые критерии (non-refoulement, Art. 2/3/8 EMRK) закреплены в §50/§45a FPG — теперь основные якоря; тезис 'работа/финансы не влияют' остаётся практикой"),
    6:  dict(gold_chunks=["AsylG-7", "AsylG-15"], secondary=[], notes="Настя 2026-06-13: §7 Abs.2 AsylG (обращение в посольство/возврат на родину = триггеры лишения статуса) + §15 Abs.1 Z5 (передача документов); нюанс 'паспорт сдавать не обязательно' = BFA-VG §39, вне корпуса"),
    7:  dict(gold_chunks=["AsylG-10"],            secondary=FPG_8_HAUPTSTUECK, notes="Настя исправила §10 AuslBG->§10 AsylG; '8. Hauptstück FPG' = глава депортации, засчитывается любой её §"),
    8:  dict(gold_chunks=[],                      secondary=[], notes="вне корпуса"),
    9:  dict(gold_chunks=["FPG-88"],              secondary=[], notes=""),
    10: dict(gold_chunks=["NAG-19"],              secondary=[], notes="известное слабое место retrieval: NAG-19 не в top-5"),
    11: dict(gold_chunks=["NAG-20", "NAG-19"],    secondary=[], notes="Настя 2026-06-13: §20 Abs.1 NAG (срок ВНЖ ограничен сроком паспорта) + §19 Abs.8 Z3 (Mängelheilung) — тот же крючок, что у q10; Указ №278/отчёт ООН = доказательства под стандарт, остаются практикой"),
    12: dict(gold_chunks=["StbG-20"],             secondary=[], notes=""),
    13: dict(gold_chunks=["NAG-DV-7", "FPG-16"],  secondary=[], notes="NAG-DV-7 ранг 8 (top-5 промах), FPG-16 ранг 70 — трудный для retrieval"),
}


def norm(s):
    return str(s or "").replace("\n", " ").strip()


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for i in range(2, 15):  # xlsx rows 2..14 == questions 1..13
        q = norm(ws.cell(i, 1).value)
        ans = norm(ws.cell(i, 2).value)
        note = norm(ws.cell(i, 3).value)
        diff = norm(ws.cell(i, 4).value).lower()
        if not q:
            continue
        n = i - 1
        m = MAPPING[n]
        in_corpus = bool(m["gold_chunks"])
        rows.append({
            "id": f"q{n:02d}",
            "difficulty": diff if diff in ("easy", "difficult") else "unknown",
            "in_corpus": in_corpus,
            "expected_behavior": "answer" if in_corpus else "refuse",
            "question": q,
            "gold_label": note,
            "gold_chunks": m["gold_chunks"],
            "gold_chunks_secondary": m["secondary"],
            "reference_answer": ans,
            "notes": m["notes"],
        })

    with OUT.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    n_in = sum(r["in_corpus"] for r in rows)
    print(f"wrote {len(rows)} questions -> {OUT}")
    print(f"  in_corpus (answer): {n_in}   out_of_corpus (refuse): {len(rows)-n_in}")
    by_diff = {}
    for r in rows:
        by_diff[r["difficulty"]] = by_diff.get(r["difficulty"], 0) + 1
    print(f"  difficulty: {by_diff}")


if __name__ == "__main__":
    main()
